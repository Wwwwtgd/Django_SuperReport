from Auto_LH.GB_Script.script.GbParameter import GbParameter
import os
import re
import openpyxl
from datetime import datetime
from Auto_LH.GB_Script.script.get_gb_data import get_gb_data
from Auto_LH.GB_Script.script.write_gb_info_to_excel import write_gb_info_to_excel
from Auto_LH.GB_Script.script.autoEdit_GB import *
current_time = datetime.now()
PATH_template = os.getcwd() + r'/Auto_LH/GB_Script/report_template/'


def deal_form_data(form_data):
    # 统一的信息
    project_name = form_data.get('project_name')  # 项目名称
    entrust_name = form_data.get('entrust_name')  # 委托单位
    sample_name = form_data.get('sample_name')  # 样品名称
    jl_c_text = form_data.get('jl_c_text')  # 监理单位
    jl_r_text = form_data.get('jl_r_text')  # 见证人
    qy_c_text = form_data.get('qy_c_text')  # 取样单位
    qy_r_text = form_data.get('qy_r_text')  # 取样人
    number_of_report = int(form_data.get('number_of_report'))  # 报告数量
    first_no = form_data.get('first_no')  # 首份报告编号
    # 每份钢板的信息
    lh = form_data.getlist('lh')  # 炉号输入框的值
    ph = form_data.getlist('ph')  # 批号输入框的值
    gbh = form_data.getlist('gbh')  # 钢板号输入框的值
    report_date = form_data.getlist('report_date')  # 报告日期
    detect_date_lh = form_data.getlist('detect_date_lh')  # 理化检测日期
    detect_date_hx = form_data.getlist('detect_date_hx')  # 化学检测日期
    come_date = form_data.getlist('come_date')  # 受样日期

    raw_no_type = form_data.getlist('raw_no_type')  # 原样编号类型 可能是炉号、批号、钢板号中的一个或多个 -
    chemical = form_data.getlist('chemical')  # 检测元素复选框选中的值，可能是多个元素的组合 -
    z_open = form_data.getlist('z_open')  # Z 向开口度输入框的值
    z_limit = form_data.getlist('z_status')  # Z 向性能限值输入框的值

    standard = form_data.getlist('standard')  # 验收标准下拉框选中的值
    material = form_data.getlist('material')  # 材质输入框的值
    thickness = form_data.getlist('thickness')  # 规格输入框的值
    status = form_data.getlist('status')  # 样品状态下拉框选中的值
    manufacturer = form_data.getlist('manufacturer')  # 生产厂家

    # 读取模板文件
    wb = openpyxl.load_workbook(PATH_template + 'list.xlsx')
    df_all_info = wb['总体信息']
    df_detailed = wb['钢板详细信息']

    gb_data = []  # 存放 GbParameter 对象 ，钢板全部参数

    # 从页面获取的公共信息参数，写入 excel_公共信息表
    # 再用一个for循环，依次处理每一份钢板的相关参数，创建GbParameter对象
    # 再将每个gp对象添加到excel——钢板信息表中
    # 最后，读取excel去生成报告

    dict_all_info = {"项目名称": project_name, "委托单位": entrust_name, "样品名称": sample_name, "监理单位": jl_c_text,
                     "见证人": jl_r_text, "取样单位": qy_c_text, "取样人": qy_r_text, "报告数量": number_of_report}
    header = [df_all_info.cell(row=row, column=1).value for row in range(1, df_all_info.max_row + 1)]
    insert_col_index = 2
    for row_index, row_name in enumerate(header, start=1):
        if row_name in dict_all_info:  # 确保表头存在
            df_all_info.cell(row=row_index, column=insert_col_index, value=dict_all_info[row_name])  # 写入数据

    for i in range(number_of_report):
        # 创建一个空的 GbParameter 对象
        gb_param = GbParameter()
        gb_param.report_no = first_no[:-3] + str(int(first_no[-3:]) + i).zfill(3) if first_no != '/' else '/'  # 报告编号
        gb_param.sz_no = 'SZ-GB-' + str(int(first_no[-3:]) + i).zfill(2) if first_no != '/' else '/'   # 样品编号
        gb_param.lh = lh[i] if i < len(lh) and lh[i] else '/'  # 炉号
        gb_param.ph = ph[i] if i < len(ph) and ph[i] else '/'  # 批号
        gb_param.gbh = gbh[i] if i < len(gbh) and gbh[i] else '/'  # 钢板号
        gb_param.standard_no = standard[i]  # 标准号
        gb_param.material = material[i]  # 材质
        gb_param.thickness = extract_number(thickness[i])  # 规格
        gb_param.status = status[i]  # 样品状态
        gb_param.manufacturer = manufacturer[i]  # 生产厂家
        gb_param.come_date = come_date[i]  # 受样日期
        gb_param.detect_date_lh = detect_date_lh[i]  # 理化检测日期
        gb_param.detect_date_hx = detect_date_hx[i]  # 化学检测日期
        gb_param.report_date = report_date[i]  # 报告日期
        gb_param.z_limit = z_limit[0].replace("Z", "≥") if len(z_limit) > 0 else '/'  # Z 向性能限值

        get_gb_data(gb_param)
        gb_data.append(gb_param)
        df_detailed = write_gb_info_to_excel(gb_param, df_detailed, i + 2)
    path_save = PATH_template + 'ALL_INFO' + current_time.strftime('%Y-%m-%d %H-%M-%S') + '.xlsx'
    # **保存 Excel（保持所有表不变）**
    wb.save(path_save)
    wb.close()
    auto_edit_gb(path_save)


def extract_number(s):
    match = re.search(r'\d+', s)  # \d+ 匹配一个或多个数字
    if match:
        return int(match.group())  # 返回提取的数字（转换为整数）
    else:
        return 0  # 如果没有找到数字，返回0


